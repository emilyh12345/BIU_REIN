import openpyxl

#path to the original rein file with all interactions
rein_file_given = input("enter the path of a rein file: ")

# create a rein output file called outputFile.rein to store the new rein file with only the optional interactions within the threshold
new_rein_file = "outputFile.rein"
my_writer = open(new_rein_file, "w")

#create a temporary text file of what the user will enter manually into Jupyter notebook(the path to the new rein file)
temp_text_file = "jupyterInput.txt"
my_writer2 = open(temp_text_file, "w")
my_writer2.write("./Science2014/" + new_rein_file + "\n Then Check if has Solutions")


# read in the excel chart file 
excel_file = input("enter the path of an excel file: ")
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# prompts the user to enter a number
user_number = float(input("Enter a threshold number: "))

# Open the rein input file in read mode
with open(rein_file_given, "r") as file:
    lines = file.readlines()

# write to the new rein file 3 parts: 
# 1) the section above the list of gene interactions 
# 2) the list of gene interactions but only the optional within the threshold
# 3) the section after the gene interactions (experiments and functions)

# section 1: 
#"positive" and "negative" are words we want to stop printing the first section at when we get to them
stop_words = ["positive", "negative"]
for line in lines:
    if any(word in line for word in stop_words): #once get to first line that has gene1 gene2 interaction_type(pos or neg) we will stop writing this first section to the output file
        break  
    my_writer.write(line + "\n") 

# section 2: iterate through the excel file to narrow down interactions
for row in sheet.iter_rows():
    for cell in row:
        gene1 = str(sheet.cell(row=1, column=cell.column).value)
        gene2 = str(sheet.cell(row=cell.row, column=1).value)
        if cell.data_type == "n" and cell.value is not None: 
            numeric_value = cell.value 
            #if user enters a positive number then print to file all relationships above that number and below its negative, and optional 
            if user_number > 0: 
                if numeric_value > user_number: 
                    for line in lines:
                        if f"{gene1} {gene2} positive optional" in line:
                            my_writer.write(line + "\n")
                            print (line + "\n") 
                if numeric_value < -(user_number):
                    for line in lines:
                        if f"{gene1} {gene2} negative optional" in line:
                            my_writer.write(line + "\n")
                            print (line + "\n")
            #if user enters a negative number then print to file all relationships below that number and above the abs value and optional 
            else:
                if numeric_value < user_number or numeric_value > abs(user_number):
                    for line in lines:
                        if numeric_value < 0: 
                            if f"{gene1} {gene2} negative optional" in line:
                                my_writer.write(line + "\n")
                                print (line + "\n")
                        else: 
                            if f"{gene1} {gene2} positive optional" in line:
                                my_writer.write(line + "\n")
                                print (line + "\n")

# Section 3
last_stop_word_index = -1
# Find the index of the last line containing a stop word
for i, line in enumerate(lines):
    if any(word in line for word in stop_words):
        last_stop_word_index = i
my_writer.write("\n".join(lines[last_stop_word_index + 1:])) #writes everything from the line after the last gene1 gene2 interaction_type line until the end of the file

my_writer.close()
wb.close()